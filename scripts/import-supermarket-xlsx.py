"""
Import supermarket Excel into Sanam Supabase:
- Keep website main section titles (store-data sections)
- Replace subcategories from Excel (mapped into those sections)
- Delete old products and load up to 20 products per subcategory
- Rename offers → أحدث العروض and seed 20 products there
"""
from __future__ import annotations

import hashlib
import json
import os
import random
import re
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r"c:\Users\aboda5422\Desktop\مجلد جديد\19,600 supermarket products.xlsx")
URL = os.environ["SANAM_URL"].rstrip("/")
KEY = os.environ["SANAM_SERVICE"]
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"
PER_SUB = 20
CUSTOM_SLUGS = {
    "plastics-section",
    "charcoal-gas",
    "roastery-weighed",
    "cheese-pickles-weighed",
    "cooking-tools",
    "summer-resort-goods",
}

SECTION_META = [
    ("offers", "أحدث العروض", "Latest Offers"),
    ("daily", "الاحتياجات اليومية", "Daily Essentials"),
    ("pantry", "مقاضي", "Pantry"),
    ("drinks", "المشروبات", "Beverages"),
    ("snacks", "السناكات والحلويات", "Snacks & Sweets"),
    ("health", "التغذية الصحية", "Healthy Nutrition"),
    ("makeup", "المكياج", "Makeup"),
    ("perfumes", "العطور", "Perfumes"),
    ("beauty", "الجمال", "Beauty"),
    ("home", "العناية بالمنزل", "Home Care"),
    ("electronics", "الإلكترونيات والعناية بالسيارة", "Electronics & Car Care"),
    ("baby", "العناية بالطفل", "Baby Care"),
    ("pets", "الحيوانات الأليفة", "Pets"),
    ("toys", "الألعاب", "Toys"),
    ("stationery", "القرطاسية", "Stationery"),
    ("pharmacy", "الصيدلية والتغذية الصحية", "Pharmacy & Health"),
]

MAIN_TO_SECTION = {
    "الاحتياجات اليومية والطازجة": "daily",
    "البقالة والمواد الغذائية": "pantry",
    "المشروبات": "drinks",
    "المخبوزات والحلويات": "snacks",
    "المجمدات": "pantry",
    "التنظيف والمنزل": "home",
    "العناية الشخصية والجمال": "beauty",
    "الأم والطفل": "baby",
    "الحيوانات الأليفة": "pets",
    "المنزل والمطبخ": "home",
    "القرطاسية والمدرسة": "stationery",
    "الإلكترونيات والجوالات": "electronics",
    "الألعاب والهوايات": "toys",
    "الرحلات والحديقة والشواء": "home",
    "السيارات": "electronics",
    "الأزياء والإكسسوارات": "home",
}

COLORS = [
    "bg-red-50",
    "bg-green-50",
    "bg-orange-50",
    "bg-blue-50",
    "bg-amber-50",
    "bg-yellow-50",
    "bg-pink-50",
    "bg-purple-50",
    "bg-cyan-50",
    "bg-teal-50",
    "bg-rose-50",
    "bg-gray-50",
]


def http(method: str, path: str, data=None, prefer: str | None = None):
    headers = {
        "apikey": KEY,
        "Authorization": f"Bearer {KEY}",
        "Content-Type": "application/json",
        "User-Agent": UA,
    }
    if prefer:
        headers["Prefer"] = prefer
    body = None if data is None else json.dumps(data, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(f"{URL}{path}", data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            raw = resp.read().decode("utf-8")
            return resp.status, json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        err = e.read().decode("utf-8", errors="replace")
        raise SystemExit(f"HTTP {e.code} {path}: {err[:800]}") from e


def slugify(en: str, ar: str, section: str) -> str:
    base = (en or ar or "cat").lower()
    base = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
    if not base or len(base) < 2:
        base = "c-" + hashlib.md5((ar or en).encode("utf-8")).hexdigest()[:8]
    return f"{section}-{base}"[:72]


def resolve_section(main_ar: str, sub_ar: str) -> str | None:
    main = (main_ar or "").strip()
    sub = (sub_ar or "").strip()
    if main == "العناية الشخصية والجمال":
        if "عطر" in sub:
            return "perfumes"
        if "مكياج" in sub or "تجميل" in sub:
            return "makeup"
        return "beauty"
    if main == "البقالة والمواد الغذائية" and ("صحي" in sub or "عضوي" in sub):
        return "health"
    return MAIN_TO_SECTION.get(main)


def parse_gallery(raw) -> list[str]:
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    parts = re.split(r"[\n,|]+", s)
    return [p.strip() for p in parts if p.strip().startswith("http")]


def build_description(brand, origin, size, form, extra) -> str:
    bits = []
    if brand:
        bits.append(f"الماركة: {brand}")
    if origin:
        bits.append(f"بلد المنشأ: {origin}")
    if size:
        bits.append(f"الحجم: {size}")
    if form:
        bits.append(f"الشكل: {form}")
    if extra:
        bits.append(str(extra))
    return "\n".join(bits) if bits else ""


def unit_from(size, form) -> str:
    s = (size or form or "قطعة").strip()
    return s[:40] if s else "قطعة"


def main() -> None:
    random.seed(42)
    if not XLSX.exists():
        raise SystemExit(f"Excel not found: {XLSX}")

    print("Reading", XLSX)
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["منتجات"]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header

    # key: (section, sub_ar) -> meta + products list
    buckets: dict[tuple[str, str], dict] = {}
    skipped = 0
    for row in rows:
        if not row or not row[1]:
            continue
        name_ar = str(row[1]).strip()
        name_en = str(row[2]).strip() if row[2] else ""
        main_ar = str(row[3]).strip() if row[3] else ""
        main_en = str(row[4]).strip() if row[4] else ""
        sub_ar = str(row[5]).strip() if row[5] else ""
        sub_en = str(row[6]).strip() if row[6] else ""
        brand = str(row[7]).strip() if row[7] else ""
        origin = str(row[8]).strip() if row[8] else ""
        size = str(row[9]).strip() if row[9] else ""
        form = str(row[10]).strip() if row[10] else ""
        try:
            price = float(row[11]) if row[11] is not None else 0
        except (TypeError, ValueError):
            price = 0
        image = str(row[12]).strip() if row[12] else ""
        gallery = parse_gallery(row[13] if len(row) > 13 else None)
        extra = str(row[14]).strip() if len(row) > 14 and row[14] else ""
        barcode = str(row[0]).strip() if row[0] else ""

        section = resolve_section(main_ar, sub_ar)
        if not section or not sub_ar or price <= 0:
            skipped += 1
            continue
        if not image:
            image = "/placeholder.png"

        key = (section, sub_ar)
        if key not in buckets:
            buckets[key] = {
                "section": section,
                "name": sub_ar,
                "name_en": sub_en or sub_ar,
                "main_ar": main_ar,
                "main_en": main_en,
                "products": [],
            }
        buckets[key]["products"].append(
            {
                "name": name_ar,
                "name_en": name_en or None,
                "price": round(price, 2),
                "image": image,
                "barcode": barcode or None,
                "brand": brand or None,
                "origin_country": origin or None,
                "size_label": size or None,
                "product_form": form or None,
                "gallery_urls": gallery,
                "extra_label": extra or None,
                "unit": unit_from(size, form),
                "description": build_description(brand, origin, size, form, extra) or None,
            }
        )
    wb.close()
    print("buckets", len(buckets), "skipped", skipped)

    # Sample PER_SUB products per subcategory
    for b in buckets.values():
        prods = b["products"]
        if len(prods) > PER_SUB:
            b["products"] = random.sample(prods, PER_SUB)
        # ensure first product has image for category tile
        imgs = [p["image"] for p in b["products"] if p["image"] and not p["image"].endswith("placeholder.png")]
        b["image"] = imgs[0] if imgs else "/placeholder.png"

    # Latest offers: diversify 20 products from all buckets
    all_pool = []
    for b in buckets.values():
        all_pool.extend(b["products"])
    offer_products = random.sample(all_pool, min(PER_SUB, len(all_pool))) if all_pool else []
    for i, p in enumerate(offer_products):
        # light "new offer" markup: some with original_price
        if i % 3 == 0:
            p = dict(p)
            p["original_price"] = round(p["price"] * 1.15, 2)
            p["is_featured"] = True
            offer_products[i] = p
        else:
            p = dict(p)
            p["is_featured"] = i < 5
            offer_products[i] = p

    offers_bucket = {
        "section": "offers",
        "name": "أحدث العروض",
        "name_en": "Latest Offers",
        "slug": "latest-offers",
        "image": offer_products[0]["image"] if offer_products else "/placeholder.png",
        "products": offer_products,
    }

    # Build category rows with stable slugs
    used_slugs: set[str] = set()
    category_defs = []
    sort_i = 1
    # offers first
    category_defs.append(
        {
            "slug": "latest-offers",
            "name": offers_bucket["name"],
            "name_en": offers_bucket["name_en"],
            "image": offers_bucket["image"],
            "section": "offers",
            "sort_order": sort_i,
            "is_active": True,
            "_products": offers_bucket["products"],
        }
    )
    used_slugs.add("latest-offers")
    sort_i += 1

    # order by SECTION_META then name
    section_order = {sid: i for i, (sid, _, _) in enumerate(SECTION_META)}
    ordered_keys = sorted(
        buckets.keys(),
        key=lambda k: (section_order.get(k[0], 99), k[1]),
    )
    for section, sub_ar in ordered_keys:
        b = buckets[(section, sub_ar)]
        slug = slugify(b["name_en"], b["name"], section)
        if slug in used_slugs or slug in CUSTOM_SLUGS:
            slug = f"{slug}-{hashlib.md5(sub_ar.encode()).hexdigest()[:4]}"
        used_slugs.add(slug)
        category_defs.append(
            {
                "slug": slug,
                "name": b["name"],
                "name_en": b["name_en"],
                "image": b["image"],
                "section": section,
                "sort_order": sort_i,
                "is_active": True,
                "_products": b["products"],
            }
        )
        sort_i += 1

    print("categories to upsert", len(category_defs))
    print("products to insert", sum(len(c["_products"]) for c in category_defs))

    # ---- DB: delete products in batches ----
    print("Deleting existing products...")
    while True:
        status, rows = http(
            "GET",
            "/rest/v1/products?select=id&limit=500",
            prefer="return=representation",
        )
        if not rows:
            break
        ids = [r["id"] for r in rows]
        # delete by id list
        id_list = ",".join(ids)
        # PostgREST: DELETE with in.()
        http("DELETE", f"/rest/v1/products?id=in.({id_list})", prefer="return=minimal")
        print("  deleted", len(ids))
        time.sleep(0.2)

    # Delete non-custom categories
    print("Deleting old categories (keep Sanam customs)...")
    status, cats = http("GET", "/rest/v1/categories?select=id,slug&limit=1000")
    to_del = [c["id"] for c in (cats or []) if c["slug"] not in CUSTOM_SLUGS]
    for i in range(0, len(to_del), 100):
        chunk = to_del[i : i + 100]
        id_list = ",".join(chunk)
        http("DELETE", f"/rest/v1/categories?id=in.({id_list})", prefer="return=minimal")
        print("  deleted cats", len(chunk))

    # Upsert custom section tags
    customs = [
        ("plastics-section", "قسم البلاستيكات", "Plastics", "home", 200),
        ("charcoal-gas", "قسم الفحم والغاز", "Charcoal & Gas", "home", 201),
        ("roastery-weighed", "قسم المحمصة (بالميزان)", "Roastery (By Weight)", "pantry", 202),
        ("cheese-pickles-weighed", "قسم الأجبان والمخللات (بالميزان)", "Cheese & Pickles (By Weight)", "daily", 203),
        ("cooking-tools", "قسم أدوات الطهي", "Cooking Tools", "home", 204),
        ("summer-resort-goods", "قسم خردوات المصيف", "Summer Resort Goods", "home", 205),
    ]
    http(
        "POST",
        "/rest/v1/categories?on_conflict=slug",
        [
            {
                "slug": s,
                "name": n,
                "name_en": ne,
                "section": sec,
                "sort_order": so,
                "is_active": True,
            }
            for s, n, ne, sec, so in customs
        ],
        prefer="resolution=merge-duplicates,return=minimal",
    )

    # Insert categories
    print("Inserting categories...")
    cat_payload = [
        {k: v for k, v in c.items() if k != "_products"} for c in category_defs
    ]
    for i in range(0, len(cat_payload), 50):
        http(
            "POST",
            "/rest/v1/categories?on_conflict=slug",
            cat_payload[i : i + 50],
            prefer="resolution=merge-duplicates,return=minimal",
        )
        print("  cats batch", i // 50 + 1)

    # Fetch category ids by slug
    status, db_cats = http(
        "GET",
        "/rest/v1/categories?select=id,slug&limit=1000",
    )
    id_by_slug = {c["slug"]: c["id"] for c in (db_cats or [])}

    # Insert products
    print("Inserting products...")
    total = 0
    batch = []
    for cdef in category_defs:
        cid = id_by_slug.get(cdef["slug"])
        if not cid:
            print("MISSING CAT", cdef["slug"])
            continue
        for idx, p in enumerate(cdef["_products"]):
            row = {
                "name": p["name"],
                "name_en": p["name_en"],
                "price": p["price"],
                "original_price": p.get("original_price"),
                "image": p["image"],
                "category_id": cid,
                "unit": p["unit"],
                "description": p["description"],
                "barcode": p["barcode"],
                "brand": p["brand"],
                "origin_country": p["origin_country"],
                "size_label": p["size_label"],
                "product_form": p["product_form"],
                "gallery_urls": p["gallery_urls"] or [],
                "extra_label": p["extra_label"],
                "is_active": True,
                "is_featured": bool(p.get("is_featured")),
                "sort_order": idx,
                "stock_quantity": 50,
            }
            batch.append(row)
            if len(batch) >= 80:
                http("POST", "/rest/v1/products", batch, prefer="return=minimal")
                total += len(batch)
                print("  products", total)
                batch = []
                time.sleep(0.15)
    if batch:
        http("POST", "/rest/v1/products", batch, prefer="return=minimal")
        total += len(batch)
        print("  products", total)

    # Write store-data categories fragment
    out_cats = []
    for i, c in enumerate(category_defs):
        color = COLORS[i % len(COLORS)]
        out_cats.append(
            {
                "id": c["slug"],
                "name": c["name"],
                "nameEn": c["name_en"],
                "icon": "🛒",
                "image": c["image"],
                "color": color,
                "section": c["section"],
            }
        )
    out_path = ROOT / "scripts" / "generated-categories.json"
    out_path.write_text(json.dumps(out_cats, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", out_path, "count", len(out_cats))
    print("DONE")


if __name__ == "__main__":
    main()
