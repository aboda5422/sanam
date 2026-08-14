"""Add remaining Excel products to every branch (no 20-per-sub sample, no deletes)."""
from __future__ import annotations

import hashlib
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r"c:\Users\aboda5422\Desktop\مجلد جديد\19,600 supermarket products.xlsx")
URL = os.environ["SANAM_URL"].rstrip("/")
KEY = os.environ["SANAM_SERVICE"]
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"
CUSTOM_SLUGS = {
    "plastics-section",
    "charcoal-gas",
    "roastery-weighed",
    "cheese-pickles-weighed",
    "cooking-tools",
    "summer-resort-goods",
}

MAIN_TO_SECTION = {
    "الاحتياجات اليومية والطازجة": "daily",
    "البقالة والمواد الغذائية": "pantry",
    "المشروبات": "drinks",
    "المخبوزات والحلويات": "snacks",
    "المجمدات": "pantry",
    "التنظيف والمنزل": "home",
    "العناية الشخصية والجمال": "beauty",
    "الأم والطفل": "baby",
    "الحيوانات الأليفة": "pets",
    "المنزل والمطبخ": "home",
    "القرطاسية والمدرسة": "stationery",
    "الإلكترونيات والجوالات": "electronics",
    "الألعاب والهوايات": "toys",
    "الرحلات والحديقة والشواء": "home",
    "السيارات": "electronics",
    "الأزياء والإكسسوارات": "home",
}


def http(method: str, path: str, data=None, prefer: str | None = None, extra_headers: dict | None = None):
    headers = {
        "apikey": KEY,
        "Authorization": f"Bearer {KEY}",
        "Content-Type": "application/json",
        "User-Agent": UA,
    }
    if prefer:
        headers["Prefer"] = prefer
    if extra_headers:
        headers.update(extra_headers)
    body = None if data is None else json.dumps(data, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(f"{URL}{path}", data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            raw = resp.read().decode("utf-8")
            return resp.status, json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        err = e.read().decode("utf-8", errors="replace")
        raise SystemExit(f"HTTP {e.code} {path}: {err[:800]}") from e


def get_all(path: str, page=1000):
    rows = []
    offset = 0
    while True:
        sep = "&" if "?" in path else "?"
        _, chunk = http("GET", f"{path}{sep}limit={page}&offset={offset}")
        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < page:
            break
        offset += page
    return rows


def slugify(en: str, ar: str, section: str) -> str:
    base = (en or ar or "cat").lower()
    base = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
    if not base or len(base) < 2:
        base = "c-" + hashlib.md5((ar or en).encode("utf-8")).hexdigest()[:8]
    return f"{section}-{base}"[:72]


def resolve_section(main_ar: str, sub_ar: str) -> str | None:
    main = (main_ar or "").strip()
    sub = (sub_ar or "").strip()
    if main == "العناية الشخصية والجمال":
        if "عطر" in sub:
            return "perfumes"
        if "مكياج" in sub or "تجميل" in sub:
            return "makeup"
        return "beauty"
    if main == "البقالة والمواد الغذائية" and ("صحي" in sub or "عضوي" in sub):
        return "health"
    return MAIN_TO_SECTION.get(main)


def parse_gallery(raw) -> list[str]:
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    parts = re.split(r"[\n,|]+", s)
    return [p.strip() for p in parts if p.strip().startswith("http")]


def build_description(brand, origin, size, form, extra) -> str:
    bits = []
    if brand:
        bits.append(f"الماركة: {brand}")
    if origin:
        bits.append(f"بلد المنشأ: {origin}")
    if size:
        bits.append(f"الحجم: {size}")
    if form:
        bits.append(f"الشكل: {form}")
    if extra:
        bits.append(str(extra))
    return "\n".join(bits) if bits else ""


def unit_from(size, form) -> str:
    s = (size or form or "قطعة").strip()
    return s[:40] if s else "قطعة"


def parse_excel():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["منتجات"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    buckets: dict[tuple[str, str], dict] = {}
    skipped = defaultdict(int)
    total = 0
    for row in rows:
        total += 1
        if not row or not row[1]:
            skipped["no_name"] += 1
            continue
        name_ar = str(row[1]).strip()
        name_en = str(row[2]).strip() if row[2] else ""
        main_ar = str(row[3]).strip() if row[3] else ""
        sub_ar = str(row[5]).strip() if row[5] else ""
        sub_en = str(row[6]).strip() if row[6] else ""
        brand = str(row[7]).strip() if row[7] else ""
        origin = str(row[8]).strip() if row[8] else ""
        size = str(row[9]).strip() if row[9] else ""
        form = str(row[10]).strip() if row[10] else ""
        try:
            price = float(row[11]) if row[11] is not None else 0
        except (TypeError, ValueError):
            skipped["bad_price"] += 1
            continue
        image = str(row[12]).strip() if row[12] else ""
        gallery = parse_gallery(row[13] if len(row) > 13 else None)
        extra = str(row[14]).strip() if len(row) > 14 and row[14] else ""
        barcode = str(row[0]).strip() if row[0] else ""
        section = resolve_section(main_ar, sub_ar)
        if not section:
            skipped["no_section"] += 1
            continue
        if not sub_ar:
            skipped["no_sub"] += 1
            continue
        if price < 0:
            skipped["neg_price"] += 1
            continue
        if not image:
            image = "/placeholder.png"
        key = (section, sub_ar)
        if key not in buckets:
            buckets[key] = {
                "section": section,
                "name": sub_ar,
                "name_en": sub_en or sub_ar,
                "products": [],
            }
        buckets[key]["products"].append(
            {
                "name": name_ar,
                "name_en": name_en or None,
                "price": round(price, 2),
                "image": image,
                "barcode": barcode or None,
                "brand": brand or None,
                "origin_country": origin or None,
                "size_label": size or None,
                "product_form": form or None,
                "gallery_urls": gallery,
                "extra_label": extra or None,
                "unit": unit_from(size, form),
                "description": build_description(brand, origin, size, form, extra) or None,
            }
        )
    wb.close()
    return buckets, dict(skipped), total


def category_defs(buckets: dict) -> list[dict]:
    used: set[str] = set()
    defs = []
    sort_i = 10
    for (section, sub_ar) in sorted(buckets.keys(), key=lambda k: (k[0], k[1])):
        b = buckets[(section, sub_ar)]
        slug = slugify(b["name_en"], b["name"], section)
        if slug in used or slug in CUSTOM_SLUGS:
            slug = f"{slug}-{hashlib.md5(sub_ar.encode()).hexdigest()[:4]}"
        used.add(slug)
        imgs = [p["image"] for p in b["products"] if p["image"] and not str(p["image"]).endswith("placeholder.png")]
        defs.append(
            {
                "slug": slug,
                "name": b["name"],
                "name_en": b["name_en"],
                "image": imgs[0] if imgs else "/placeholder.png",
                "section": section,
                "sort_order": sort_i,
                "is_active": True,
                "_products": b["products"],
            }
        )
        sort_i += 1
    return defs


def existing_keys(branch_id: str) -> tuple[set[str], set[tuple[str, str]]]:
    barcodes: set[str] = set()
    name_cat: set[tuple[str, str]] = set()
    rows = get_all(f"/rest/v1/products?select=barcode,name,category_id&branch_id=eq.{branch_id}")
    for r in rows:
        if r.get("barcode"):
            barcodes.add(str(r["barcode"]).strip())
        if r.get("name") and r.get("category_id"):
            name_cat.add((r["name"], r["category_id"]))
    return barcodes, name_cat


def ensure_categories(branch_id: str, defs: list[dict]) -> dict[str, str]:
    existing = get_all(f"/rest/v1/categories?select=id,slug&branch_id=eq.{branch_id}")
    id_by_slug = {c["slug"]: c["id"] for c in existing}
    missing = []
    for d in defs:
        if d["slug"] not in id_by_slug:
            missing.append(
                {
                    "slug": d["slug"],
                    "name": d["name"],
                    "name_en": d["name_en"],
                    "image": d["image"],
                    "section": d["section"],
                    "sort_order": d["sort_order"],
                    "is_active": True,
                    "branch_id": branch_id,
                }
            )
    for i in range(0, len(missing), 50):
        http("POST", "/rest/v1/categories", missing[i : i + 50], prefer="return=minimal")
        print(f"    new cats {min(i+50, len(missing))}/{len(missing)}")
        time.sleep(0.1)
    if missing:
        existing = get_all(f"/rest/v1/categories?select=id,slug&branch_id=eq.{branch_id}")
        id_by_slug = {c["slug"]: c["id"] for c in existing}
    return id_by_slug


def insert_for_branch(branch: dict, defs: list[dict]) -> tuple[int, int]:
    bid = branch["id"]
    print("Branch", branch.get("name"), bid)
    id_by_slug = ensure_categories(bid, defs)
    barcodes, name_cat = existing_keys(bid)
    print("  already", len(barcodes), "barcodes,", len(name_cat), "name+cat")
    batch = []
    inserted = 0
    skipped_dup = 0

    def flush():
        nonlocal batch, inserted
        if not batch:
            return
        http("POST", "/rest/v1/products", batch, prefer="return=minimal")
        inserted += len(batch)
        print("  inserted", inserted)
        batch = []
        time.sleep(0.08)

    for cdef in defs:
        cid = id_by_slug.get(cdef["slug"])
        if not cid:
            print("  MISSING CAT", cdef["slug"])
            continue
        for idx, p in enumerate(cdef["_products"]):
            bc = p["barcode"]
            if bc and bc in barcodes:
                skipped_dup += 1
                continue
            if (p["name"], cid) in name_cat:
                skipped_dup += 1
                continue
            batch.append(
                {
                    "name": p["name"],
                    "name_en": p["name_en"],
                    "price": p["price"],
                    "image": p["image"],
                    "category_id": cid,
                    "unit": p["unit"],
                    "description": p["description"],
                    "barcode": p["barcode"],
                    "brand": p["brand"],
                    "origin_country": p["origin_country"],
                    "size_label": p["size_label"],
                    "product_form": p["product_form"],
                    "gallery_urls": p["gallery_urls"] or [],
                    "extra_label": p["extra_label"],
                    "is_active": True,
                    "is_featured": False,
                    "sort_order": idx,
                    "stock_quantity": 50,
                    "branch_id": bid,
                }
            )
            if bc:
                barcodes.add(bc)
            name_cat.add((p["name"], cid))
            if len(batch) >= 80:
                flush()
    flush()
    return inserted, skipped_dup


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Excel not found: {XLSX}")
    print("Reading", XLSX)
    buckets, skipped, total_rows = parse_excel()
    defs = category_defs(buckets)
    n_prods = sum(len(c["_products"]) for c in defs)
    print("excel rows", total_rows, "skipped", skipped)
    print("importable products", n_prods, "subcategories", len(defs))

    branches = get_all("/rest/v1/branches?select=id,name,slug&is_active=eq.true")
    print("branches", [(b["name"], b["id"]) for b in branches])
    if not branches:
        raise SystemExit("No branches")

    for b in branches:
        ins, dup = insert_for_branch(b, defs)
        print("  done inserted", ins, "already_present", dup)

    print("DONE")


if __name__ == "__main__":
    main()
